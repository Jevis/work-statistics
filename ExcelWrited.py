import openpyxl
import datetime
import math


class ExcelWrited():
    def __init__(self, name, result):
        self.name = name
        self.excel = openpyxl.load_workbook(name)
        self.excel_name = self.excel.sheetnames
        self.sheet = self.excel.get_sheet_by_name(self.excel_name[0])  # 获取第一个sheet
        self.result = result
        self.JBDuring = 2
        self.CList = []  # 迟到的集合
        self.ZList = []  # 早退的集合
        self.JList = []  # 加班的集合
        self.C_col = 7
        self.Z_col = 8
        self.J_F_col = 17
        self.J_S_col = 18

    def toWriteData(self):
        for i in range(len(self.result)):
            onePeopleData = self.result[i]
            for j in range(len(onePeopleData)):
                timeData = onePeopleData[j]
                if 1 == timeData['type']:
                    # 上午
                    mRowNum = timeData['rowNum']
                    uRT = datetime.datetime.strptime(timeData['rulerTime'], '%Y/%m/%d %H:%M')
                    uWT = datetime.datetime.strptime(timeData['workedTime'], '%Y/%m/%d %H:%M')
                    if (uWT.timestamp() - uRT.timestamp()) > 0:
                        t = int((uWT.timestamp() - uRT.timestamp()) / 60)
                        self.CList.append({'rowNumber': mRowNum, 'time': str(t)})

                else:
                    # 下午
                    dRT = datetime.datetime.strptime(timeData['rulerTime'], '%Y/%m/%d %H:%M')
                    dWT = datetime.datetime.strptime(timeData['workedTime'], '%Y/%m/%d %H:%M')
                    if (dWT.timestamp() - dRT.timestamp()) < 0:
                        u = int((dRT.timestamp() - dWT.timestamp()) / 60)
                        self.ZList.append({'rowNumber': mRowNum, 'time': str(u)})

                    else:
                        if (dWT.timestamp() - dRT.timestamp()) > (60 * 60):
                            v = (dWT.timestamp() - dRT.timestamp()) / 3600.0 / 0.5
                            if v % 2 == 0:
                                v = int(v)
                            self.JList.append({'rowNumber': mRowNum, 'time': str(
                                math.floor(v / 2))})

    def toWriteExcel(self):
        for i in range(len(self.CList)):
            CData = self.CList[i]
            self.sheet.cell(int(CData['rowNumber']) + 1, self.C_col, str(CData['time']))

        for i in range(len(self.ZList)):
            ZData = self.ZList[i]
            self.sheet.cell(int(ZData['rowNumber']) + 1, self.Z_col, str(ZData['time']))

        for i in range(len(self.JList)):
            JData = self.JList[i]
            ff = (float(JData['time']) / 0.5) % 2 == 0
            ft = float(JData['time'])
            if ft <= self.JBDuring:
                self.sheet.cell(int(JData['rowNumber']) + 1, self.J_F_col, str(JData['time']))
            else:
                self.sheet.cell(int(JData['rowNumber']) + 1, self.J_F_col, str(2))
                if ff:
                    self.sheet.cell(int(JData['rowNumber']) + 1, self.J_S_col, str(int(JData['time']) - self.JBDuring))
                else:
                    self.sheet.cell(int(JData['rowNumber']) + 1, self.J_S_col, str(ft - self.JBDuring))

        self.excel.save(self.name)
